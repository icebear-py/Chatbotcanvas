from PyPDF2 import PdfReader
from openpyxl import load_workbook
from docx import Document
import io


def extract_text_from_pdf(file_bytes: bytes):
    reader = PdfReader(io.BytesIO(file_bytes))
    text = ""
    for page in reader.pages:
        text += page.extract_text() + "\n"
    return text


def extract_text_from_docx(file_bytes: bytes):
    doc = Document(io.BytesIO(file_bytes))
    text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
    return text


def extract_txt_from_txt(file_bytes: bytes):
    return file_bytes.decode("utf-8", errors="ignore")


def extract_text_from_excel(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_text = " ".join([str(cell) for cell in row if cell is not None])
            if row_text.strip():
                text += row_text + "\n"
    return text

print(extract_text_from_url("https://www.bbotclothing.com/faqs"))


